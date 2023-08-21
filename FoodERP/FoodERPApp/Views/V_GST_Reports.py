from decimal import Decimal
from django.http import JsonResponse
from rest_framework.generics import CreateAPIView
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from rest_framework.parsers import JSONParser
from ..Views.V_TransactionNumberfun import GetMaxNumber, GetPrifix

from ..Serializer.S_GST_Reports import *
 
from ..models import  *
import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
import pandas as pd

class AllGSTReportsDownloadView(CreateAPIView):
    permission_classes = (IsAuthenticated,)
    
    def post(self, request):
        try:
            with transaction.atomic():
                Orderdata = JSONParser().parse(request)
                FromDate = Orderdata['FromDate']
                ToDate = Orderdata['ToDate']
                Party = Orderdata['Party']
                
                # Create a new workbook
                wb = Workbook()
                ws = wb.active
                ws.title = 'GST Wise'
                GSTWisequery = T_Invoices.objects.raw('''SELECT 1 as id, TC_InvoiceItems.GSTPercentage,SUM(BasicAmount)TaxableValue,SUM(CGST)CGST, SUM(SGST)SGST, SUM(IGST)IGST,SUM(CGST)+SUM(SGST)+ SUM(IGST)GSTAmount,SUM(Amount)TotalValue FROM TC_InvoiceItems JOIN T_Invoices ON TC_InvoiceItems.Invoice_id=T_Invoices.id WHERE T_Invoices.Party_id=%s  AND  T_Invoices.InvoiceDate BETWEEN %s AND %s  GROUP BY TC_InvoiceItems.GSTPercentage UNION SELECT 1 as id,0 as GSTPercentage,0 TaxableValue,0 CGST, 0 SGST, 0 IGST,0 GSTAmount,SUM(T_Invoices.TCSAmount)TotalValue FROM T_Invoices WHERE T_Invoices.Party_id=%s  AND  T_Invoices.InvoiceDate BETWEEN %s AND %s ''',([Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate]))
               
                GSTWisedata = GSTWiseSerializer(GSTWisequery, many=True).data
                df=pd.DataFrame(GSTWisedata)
                if not df.empty:

                    specific_column_names = {
                    'GSTPercentage':'GST Rate',
                    'TaxableValue':'Taxable Value', 
                    'CGST':'CGST',
                    'SGST':'SGST',
                    'IGST':'IGST',
                    'GSTAmount':'GST Amount',
                    'TotalValue' :'Total Value',
                    
                    }
                
                    # Define which columns header Font bold
                    for col_idx, header in enumerate(df.columns, start=1):
                        
                        cell = ws.cell(row=2, column=col_idx, value=specific_column_names.get(header, header))
                        bold_font = Font(bold=True)
                        cell.font = bold_font
                        
                    # Write the data
                    for col_idx, header in enumerate(df.columns, start=1):
                        for row_idx, value in enumerate(df[header], start=3):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    max_cols = len(df.columns)

                    # Merge cells and add the hardcoded text
                   
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
                    merged_cell = ws.cell(row=1, column=1, value="GST Wise Summary")
                    bold_font = Font(bold=True)
                    merged_cell.font = bold_font
                    merged_cell.alignment = Alignment(horizontal='center')  # Align text to center
                
                    
                    B2Bquery2 = T_Invoices.objects.raw('''SELECT 'Total' as id, SUM(TaxableValue)TotalTaxableValue,SUM(CGST)TotalCGST,SUM(SGST)TotalSGST,SUM(IGST)TotalIGST,SUM(GSTAmount)TotalGSTAmount,SUM(TotalValue)GrandTotal FROM  (SELECT TC_InvoiceItems.GSTPercentage,SUM(BasicAmount)TaxableValue,SUM(CGST)CGST, SUM(SGST)SGST, SUM(IGST)IGST,SUM(CGST)+SUM(SGST)+ SUM(IGST)GSTAmount,SUM(Amount)TotalValue FROM TC_InvoiceItems JOIN T_Invoices ON TC_InvoiceItems.Invoice_id=T_Invoices.id WHERE T_Invoices.Party_id=%s  AND  T_Invoices.InvoiceDate BETWEEN %s AND %s  GROUP BY TC_InvoiceItems.GSTPercentage UNION SELECT  0 GSTPercentage,0 TaxableValue,0 CGST, 0 SGST, 0 IGST,0 GSTAmount,SUM(T_Invoices.TCSAmount)TotalValue FROM T_Invoices WHERE T_Invoices.Party_id=%s  AND  T_Invoices.InvoiceDate BETWEEN %s AND %s )a  ''',([Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate]))
                    GSTWisedata2 = GSTWiseSerializer2(B2Bquery2, many=True).data
                    GSTWisedf2=pd.DataFrame(GSTWisedata2)
                    # print(B2Bdf2)
                    
                    specific_column_names = {
                    'id':'id',
                    'TotalTaxableValue':'TotalTaxableValue',
                    'TotalCGST':'TotalCGST',
                    'TotalSGST':'TotalSGST',
                    'TotalIGST':'TotalIGST',
                    'TotalGSTAmount':'TotalGSTAmount',
                    'GrandTotal':'GrandTotal',
                    }
                
                    start_row_idx_df2 = len(df) + 4  # Add a gap of 5 rows between the two dataframes

                    # Write the second dataframe's data
                    for col_idx, header in enumerate(GSTWisedf2.columns, start=1):
                        for row_idx, value in enumerate(GSTWisedf2[header], start=start_row_idx_df2):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                                

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=example.xlsx'

                return response
        except Exception as e:
            return JsonResponse({'StatusCode': 400, 'Status': True, 'Message':  Exception(e), 'Data': []})

