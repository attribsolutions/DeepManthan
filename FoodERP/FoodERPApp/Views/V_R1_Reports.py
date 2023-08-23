from decimal import Decimal
from django.http import JsonResponse
from rest_framework.generics import CreateAPIView
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from rest_framework.parsers import JSONParser
from ..Views.V_TransactionNumberfun import GetMaxNumber, GetPrifix

from ..Serializer.S_R1_Reports import *
 
from ..models import  *
import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
import pandas as pd

class GSTR1ExcelDownloadView(CreateAPIView):
    permission_classes = (IsAuthenticated,)
    
    def post(self, request):
        try:
            with transaction.atomic():
                Orderdata = JSONParser().parse(request)
                FromDate = Orderdata['FromDate']
                ToDate = Orderdata['ToDate']
                Party = Orderdata['Party']
                
                # Create a new workbook
                wb = Workbook()
                ws = wb.active
                ws.title = 'B2B'
                B2Bquery = T_Invoices.objects.raw('''SELECT T_Invoices.id, M_Parties.GSTIN , M_Parties.Name,T_Invoices.FullInvoiceNumber,T_Invoices.InvoiceDate,(T_Invoices.GrandTotal+T_Invoices.TCSAmount)GrandTotal,
        concat(M_States.StateCode,'-',M_States.Name)aa, 'N' ReverseCharge,'' ApplicableofTaxRate ,'Reguler' InvoiceType,'' ECommerceGSTIN,TC_InvoiceItems.GSTPercentage Rate,sum(TC_InvoiceItems.BasicAmount) TaxableValue ,'0' CessAmount
        FROM T_Invoices 
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties ON M_Parties.id=T_Invoices.Customer_id
        JOIN M_States ON M_States.id=M_Parties.State_id
        WHERE Party_id=%s and InvoiceDate BETWEEN %s AND %s and M_Parties.GSTIN !=''
        Group by M_Parties.GSTIN,M_Parties.Name,T_Invoices.id,T_Invoices.InvoiceDate,M_States.id,M_States.Name,TC_InvoiceItems.GSTPercentage''',([Party],[FromDate],[ToDate]))
                B2Bdata = B2BSerializer(B2Bquery, many=True).data
                df=pd.DataFrame(B2Bdata)
                # print (df)
                
                specific_column_names = {
                'GSTIN':'GSTIN/UIN of Recipient', 
                'Name':'Receiver Name',
                'FullInvoiceNumber':'Invoice Number',
                'InvoiceDate':'Invoice date',
                'GrandTotal':'Invoice Value',
                'aa':'Place Of Supply',
                'ReverseCharge':'Reverse Charge',
                'ApplicableofTaxRate':'Applicable %'+'of TaxRate',
                'InvoiceType' :'Invoice Type',
                'ECommerceGSTIN':'E-Commerce GSTIN',
                'Rate':'Rate',
                'TaxableValue' :'Taxable Value',
                'CessAmount':'Cess Amount',
                }
            
                # Define which columns header Font bold
                for col_idx, header in enumerate(specific_column_names, start=1):
                    
                    cell = ws.cell(row=4, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    cell.font = bold_font
                    

                # Write the data
                for col_idx, header in enumerate(df.columns, start=1):
                    for row_idx, value in enumerate(df[header], start=5):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                max_cols = 13

                # Merge cells and add the hardcoded text
                
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
                merged_cell = ws.cell(row=1, column=1, value="Summary For B2B(4)")
                bold_font = Font(bold=True)
                merged_cell.font = bold_font
                merged_cell.alignment = Alignment(horizontal='center')  # Align text to center
            
                
                B2Bquery2 = T_Invoices.objects.raw('''SELECT 1 as id, count(DISTINCT Customer_id) NoofRecipients,count(*)NoOfInvoices,sum(T_Invoices.GrandTotal+T_Invoices.TCSAmount) TotalInvoiceValue FROM T_Invoices JOIN M_Parties ON M_Parties.id=T_Invoices.Customer_id WHERE T_Invoices.Party_id=%s AND InvoiceDate BETWEEN %s AND %s and M_Parties.GSTIN !='' ''',([Party],[FromDate],[ToDate]))
                B2Bdata2 = B2BSerializer2(B2Bquery2, many=True).data
                B2Bdf2=pd.DataFrame(B2Bdata2)
                # print(B2Bdf2)
                
                specific_column_names = {
                'NoofRecipients':'No.of Recipients', 
                'NoOfInvoices':'No.of Invoices',
                'TotalInvoiceValue':'Total Invoice Value',
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    ws.cell(row=2, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    ws.cell(row=2, column=col_idx).font = bold_font

                for col_idx, header in enumerate(B2Bdf2.columns, start=1):
                    for row_idx, value in enumerate(B2Bdf2[header], start=3):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                                
                                
                                

                # Example data for the second sheet B2CL
                B2CLquery = T_Invoices.objects.raw('''SELECT T_Invoices.id,T_Invoices.FullInvoiceNumber,T_Invoices.InvoiceDate,(T_Invoices.GrandTotal)GrandTotal,concat(M_States.StateCode,'-',M_States.Name)aa, '' ApplicableofTaxRate ,TC_InvoiceItems.GSTPercentage Rate,sum(TC_InvoiceItems.BasicAmount) TaxableValue ,'0' CessAmount,'' ECommerceGSTIN
        FROM T_Invoices 
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id
        JOIN M_States ON M_States.id=b.State_id
        where Party_id=%s and InvoiceDate BETWEEN %s AND %s and b.GSTIN !='' and b.State_id != a.State_id and T_Invoices.GrandTotal > 250000
        group by T_Invoices.id,T_Invoices.InvoiceDate,M_States.id,M_States.Name, TC_InvoiceItems.GSTPercentage''',([Party],[FromDate],[ToDate]))
                B2CLdata = B2CLSerializer(B2CLquery, many=True).data
                df2=pd.DataFrame(B2CLdata)
                ws2 = wb.create_sheet(title="B2CL")
            
                specific_column_names = {
                'FullInvoiceNumber':'Invoice Number',
                'InvoiceDate':'Invoice date',
                'GrandTotal':'Invoice Value',
                'aa':'Place Of Supply',
                'ApplicableofTaxRate':'Applicable %'+'of TaxRate',
                'ECommerceGSTIN':'E-Commerce GSTIN',
                'Rate':'Rate',
                'TaxableValue' :'Taxable Value',
                'CessAmount':'Cess Amount',
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    cell = ws2.cell(row=4, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    cell.font = bold_font
                
                # Write the data on the second worksheet
                for col_idx, header in enumerate(df2.columns, start=1):
                    for row_idx, value in enumerate(df2[header], start=5):
                        ws2.cell(row=row_idx, column=col_idx, value=value)
                
                max_cols = 9
                
                
                ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
                merged_cell = ws2.cell(row=1, column=1, value="Summary For B2CL(5)")
                bold_font = Font(bold=True)
                merged_cell.font = bold_font
                merged_cell.alignment = Alignment(horizontal='center')  # Align text to center 
                
                B2CLquery2 = T_Invoices.objects.raw('''SELECT 1 as id, count(*)NoOfInvoices,sum(T_Invoices.GrandTotal)TotalInvoiceValue FROM T_Invoices JOIN M_Parties a ON a.id=T_Invoices.Party_id JOIN M_Parties b ON b.id=T_Invoices.Customer_id JOIN M_States ON M_States.id=b.State_id WHERE Party_id=%s and InvoiceDate BETWEEN %s AND %s and b.GSTIN !='' and b.State_id != a.State_id and T_Invoices.GrandTotal > 250000 group by T_Invoices.id''',([Party],[FromDate],[ToDate]))
                B2BCLdata2 = B2CLSerializer2(B2CLquery2, many=True).data
                B2CLdf2=pd.DataFrame(B2BCLdata2)
                # print(B2Bdf2)
                
                specific_column_names = {
                'NoOfInvoices':'No.of Invoices',
                'TotalInvoiceValue':'Total Invoice Value',
                # 'TaxableValue':'Total Invoice Taxable Value',
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    ws2.cell(row=2, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    ws2.cell(row=2, column=col_idx).font = bold_font

                for col_idx, header in enumerate(B2CLdf2.columns, start=1):
                    for row_idx, value in enumerate(B2CLdf2[header], start=3):
                        ws2.cell(row=row_idx, column=col_idx, value=value)
                    
                        
                # Example data for the third sheet B2CS       
                B2CSquery = T_Invoices.objects.raw('''SELECT 1 as id, 'OE' Type,concat(M_States.StateCode,'-',M_States.Name)aa, '' ApplicableofTaxRate ,TC_InvoiceItems.GSTPercentage Rate,sum(TC_InvoiceItems.BasicAmount) TaxableValue ,'0' CessAmount,'' ECommerceGSTIN
        from T_Invoices 
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id
        JOIN M_States ON M_States.id=b.State_id
        where Party_id=%s and InvoiceDate BETWEEN %s AND %s and  b.GSTIN =''
        and ((a.State_id = b.State_id) OR (a.State_id != b.State_id and T_Invoices.GrandTotal <= 250000))
        group  by M_States.id,M_States.Name,TC_InvoiceItems.GSTPercentage''',([Party],[FromDate],[ToDate]))
            
                B2CSdata = B2CSSerializer(B2CSquery, many=True).data
                df3=pd.DataFrame(B2CSdata)
                ws3 = wb.create_sheet(title="B2CS")
              
                specific_column_names = {
                'Type':'Type',
                'aa':'Place Of Supply',
                'ApplicableofTaxRate':'Applicable %'+'of TaxRate',
                'ECommerceGSTIN':'E-Commerce GSTIN',
                'Rate':'Rate',
                'TaxableValue' :'Taxable Value',
                'CessAmount':'Cess Amount',
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    cell = ws3.cell(row=4, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    cell.font = bold_font
                
                # Write the data on the second worksheet
                for col_idx, header in enumerate(df3.columns, start=1):
                    for row_idx, value in enumerate(df3[header], start=5):
                        ws3.cell(row=row_idx, column=col_idx, value=value)
                
                max_cols = 7                
                
                ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
                merged_cell = ws3.cell(row=1, column=1, value="Summary For B2CS(7)")
                bold_font = Font(bold=True)
                merged_cell.font = bold_font
                merged_cell.alignment = Alignment(horizontal='center')  # Align text to center
            
        
                B2CSquery2 = T_Invoices.objects.raw('''SELECT 1 as id,sum(TC_InvoiceItems.BasicAmount) TaxableValue ,'' CessAmount
        from T_Invoices
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id
        JOIN M_States ON M_States.id=b.State_id
        where Party_id=%s and InvoiceDate BETWEEN  %s AND %s and  b.GSTIN =''
        and ((a.State_id = b.State_id) OR (a.State_id != b.State_id and T_Invoices.GrandTotal <= 250000))''',([Party],[FromDate],[ToDate]))
            
                B2CSdata2 = B2CSSerializer2(B2CSquery2, many=True).data
                B2CSdf3=pd.DataFrame(B2CSdata2)
                
            
                specific_column_names = {
                'TaxableValue':'Total Taxable Value',
                'CessAmount':' Total Cess',
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    ws3.cell(row=2, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    ws3.cell(row=2, column=col_idx).font = bold_font

                for col_idx, header in enumerate(B2CSdf3.columns, start=1):
                    for row_idx, value in enumerate(B2CSdf3[header], start=3):
                        ws3.cell(row=row_idx, column=col_idx, value=value)
                
                    
                # Example data for the four sheet CDNR        
                CDNRquery = T_CreditDebitNotes.objects.raw('''SELECT T_CreditDebitNotes.id, M_Parties.GSTIN,M_Parties.Name,T_CreditDebitNotes.FullNoteNumber,T_CreditDebitNotes.CRDRNoteDate,M_GeneralMaster.Name NoteTypeName,T_CreditDebitNotes.NoteType_id,CONCAT(M_States.StateCode, '-', M_States.Name) aa,'N' ReverseCharge,'Regular' NoteSupplyType,(T_CreditDebitNotes.GrandTotal) GrandTotal,'' ApplicableofTaxRate,TC_CreditDebitNoteItems.GSTPercentage Rate,SUM(TC_CreditDebitNoteItems.BasicAmount) TaxableValue,'' CessAmount FROM T_CreditDebitNotes
                JOIN TC_CreditDebitNoteItems ON TC_CreditDebitNoteItems.CRDRNote_id = T_CreditDebitNotes.id
                JOIN M_Parties ON M_Parties.id = T_CreditDebitNotes.Customer_id
                JOIN M_States ON M_States.id = M_Parties.State_id
                JOIN M_GeneralMaster ON  M_GeneralMaster.id = T_CreditDebitNotes.NoteType_id
                WHERE T_CreditDebitNotes.Party_id = %s  AND T_CreditDebitNotes.CRDRNoteDate BETWEEN %s AND %s AND M_Parties.GSTIN != '' 
                GROUP BY T_CreditDebitNotes.id, M_Parties.GSTIN , M_Parties.Name , T_CreditDebitNotes.FullNoteNumber , T_CreditDebitNotes.CRDRNoteDate,NoteTypeName, T_CreditDebitNotes.NoteType_id , M_States.id , M_States.Name , TC_CreditDebitNoteItems.GSTPercentage''',([Party],[FromDate],[ToDate]))
             
                CDNRdata = CDNRSerializer(CDNRquery, many=True).data
                df4=pd.DataFrame(CDNRdata)
                ws4 = wb.create_sheet(title="CDNR")
               
                specific_column_names = {
                'GSTIN':'GSTIN/UIN of Recipient', 
                'Name':'Receiver Name',
                'FullNoteNumber':'Note Number',
                'CRDRNoteDate':'Note date',
                'NoteTypeName':'NoteTypeName',
                'aa':'Place Of Supply',
                'ReverseCharge':'Reverse Charge',
                'GrandTotal':'Note Value',
                'ApplicableofTaxRate':'Applicable %'+'of TaxRate',
                'Rate':'Rate',
                'TaxableValue' :'Taxable Value',
                'CessAmount':'Cess Amount',
                }
                
                for col_idx, header in enumerate(specific_column_names, start=1):
                    cell = ws4.cell(row=4, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    cell.font = bold_font
            
                # Write the data on the second worksheet
                for col_idx, header in enumerate(df4.columns, start=1):
                    for row_idx, value in enumerate(df4[header], start=5):
                        ws4.cell(row=row_idx, column=col_idx, value=value)        
                    
                max_cols = 12
                
                ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
                merged_cell = ws4.cell(row=1, column=1, value="Summary For CDNR(9B)")
                bold_font = Font(bold=True)
                merged_cell.font = bold_font
                merged_cell.alignment = Alignment(horizontal='center')  # Align text to center
            
            
                CDNRquery2= T_CreditDebitNotes.objects.raw('''SELECT 1 as id, COUNT(DISTINCT A.Customer_id)NoofRecipients,COUNT(A.CRDRNote_id) NoOfNotes,SUM(A.GrandTotal) TotalInvoiceValue,SUM(A.TaxbleAmount) TotalTaxableValue, 0 CessAmount
        FROM (
        SELECT  T_CreditDebitNotes.Customer_id,TC_CreditDebitNoteItems.CRDRNote_id,T_CreditDebitNotes.GrandTotal,SUM(TC_CreditDebitNoteItems.BasicAmount) TaxbleAmount
        FROM TC_CreditDebitNoteItems
        JOIN T_CreditDebitNotes ON T_CreditDebitNotes.id= TC_CreditDebitNoteItems.CRDRNote_id
        JOIN M_Parties ON M_Parties.id = T_CreditDebitNotes.Customer_id
        WHERE Party_id=%s and T_CreditDebitNotes.CRDRNoteDate BETWEEN  %s  AND %s AND M_Parties.GSTIN != ''  Group by T_CreditDebitNotes.id)A''',([Party],[FromDate],[ToDate]))
                
                CDNR2data = CDNRSerializer2(CDNRquery2, many=True).data
                CDNRdf4=pd.DataFrame(CDNR2data)
            
            
                specific_column_names = {
                'NoofRecipients':'No.of Recipients',
                'NoOfNotes':'No.of Notes',
                'TotalInvoiceValue':' Total Invoice Value',
                'TotalTaxableValue':' Total Taxable Value',
                'CessAmount':'Total Cess',
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    ws4.cell(row=2, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    ws4.cell(row=2, column=col_idx).font = bold_font

                for col_idx, header in enumerate(CDNRdf4.columns, start=1):
                    for row_idx, value in enumerate(CDNRdf4[header], start=3):
                        ws4.cell(row=row_idx, column=col_idx, value=value)
            
                
                
                
                
                # Example data for the five sheet CDNUR 
                CDNURquery = T_CreditDebitNotes.objects.raw('''SELECT T_CreditDebitNotes.id,'' URType, T_CreditDebitNotes.FullNoteNumber,T_CreditDebitNotes.CRDRNoteDate, (CASE WHEN T_CreditDebitNotes.NoteType_id = 37 THEN 'C' ELSE 'D' END) NoteType, CONCAT(M_States.StateCode, '-', M_States.Name) aa, (T_CreditDebitNotes.GrandTotal) GrandTotal, '' ApplicableofTaxRate,TC_CreditDebitNoteItems.GSTPercentage Rate, SUM(TC_CreditDebitNoteItems.BasicAmount) TaxableValue, '' CessAmount
                FROM T_CreditDebitNotes
                JOIN TC_CreditDebitNoteItems ON TC_CreditDebitNoteItems.CRDRNote_id = T_CreditDebitNotes.id
                JOIN M_Parties ON M_Parties.id = T_CreditDebitNotes.Customer_id
                JOIN M_States ON M_States.id = M_Parties.State_id
                WHERE T_CreditDebitNotes.Party_id = %s AND T_CreditDebitNotes.CRDRNoteDate BETWEEN %s AND %s AND M_Parties.GSTIN = ''
                GROUP BY  M_Parties.Name ,T_CreditDebitNotes.CRDRNoteDate,M_States.id ,M_States.Name ,TC_CreditDebitNoteItems.GSTPercentage''',([Party],[FromDate],[ToDate]))
                CDNURdata = CDNURSerializer(CDNURquery, many=True).data
                df5=pd.DataFrame(CDNURdata)
                ws5 = wb.create_sheet(title="CDNUR")
                    
                specific_column_names = {
                'URType':'URType', 
                'FullNoteNumber':'Note Number',
                'CRDRNoteDate':'Note date',
                'NoteType':'Note Type',
                'aa':'Place Of Supply',
                'GrandTotal':'Note Value',
                'ApplicableofTaxRate':'Applicable %'+'of TaxRate',
                'Rate':'Rate',
                'TaxableValue' :'Taxable Value',
                'CessAmount':'Cess Amount',
                }
                
                for col_idx, header in enumerate(specific_column_names, start=1):
                    cell = ws5.cell(row=4, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    cell.font = bold_font
                
                # Write the data on the second worksheet
                for col_idx, header in enumerate(df5.columns, start=1):
                    for row_idx, value in enumerate(df5[header], start=5):
                        ws5.cell(row=row_idx, column=col_idx, value=value) 
                        
                max_cols = 10
                    
                ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
                merged_cell = ws5.cell(row=1, column=1, value="Summary For CDNUR(9B)")
                bold_font = Font(bold=True)
                merged_cell.font = bold_font
                merged_cell.alignment = Alignment(horizontal='center')  # Align text to center
            
            
                
                CDNURquery2= T_CreditDebitNotes.objects.raw('''SELECT 1 as id, COUNT(DISTINCT A.Customer_id)NoofRecipients,COUNT(A.CRDRNote_id) NoOfNotes,SUM(A.GrandTotal) TotalInvoiceValue,SUM(A.TaxbleAmount)
        TotalTaxableValue, 0 CessAmount FROM (SELECT T_CreditDebitNotes.Customer_id, TC_CreditDebitNoteItems.CRDRNote_id,T_CreditDebitNotes.GrandTotal,
        SUM(TC_CreditDebitNoteItems.BasicAmount) TaxbleAmount 
        FROM TC_CreditDebitNoteItems 
        JOIN T_CreditDebitNotes ON T_CreditDebitNotes.id= TC_CreditDebitNoteItems.CRDRNote_id
        JOIN M_Parties ON M_Parties.id = T_CreditDebitNotes.Customer_id 
        WHERE Party_id=%s and T_CreditDebitNotes.CRDRNoteDate BETWEEN
        %s AND  %s AND M_Parties.GSTIN = '' Group by T_CreditDebitNotes.id)A''',([Party],[FromDate],[ToDate]))
            
                CDNUR2data = CDNURSerializer2(CDNURquery2, many=True).data
                CDNURdf5=pd.DataFrame(CDNUR2data)
            
                specific_column_names = {
                'NoOfNotes':'No.of Notes',
                'TotalInvoiceValue':' Total Note Value',
                'TotalTaxableValue':' Total Taxable Value',
                'CessAmount':'Total Cess',
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    ws5.cell(row=2, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    ws5.cell(row=2, column=col_idx).font = bold_font

                for col_idx, header in enumerate(CDNURdf5.columns, start=1):
                    for row_idx, value in enumerate(CDNURdf5[header], start=3):
                        ws5.cell(row=row_idx, column=col_idx, value=value)
            
                
                
                # Example data for the six sheet CDNUR         
                EXEMPquery = T_Invoices.objects.raw('''SELECT 1 as id , 'Inter-State supplies to registered persons' Description,sum(TC_InvoiceItems.Amount) Total
        FROM T_Invoices
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id  
        WHERE Party_id= %s and T_Invoices.InvoiceDate BETWEEN %s AND %s and b.GSTIN != '' and TC_InvoiceItems.GSTPercentage= 0  and a.State_id != b.State_id group by id,Description
        UNION
        SELECT 1 as id, 'Intra-State supplies to registered persons' Description,sum(TC_InvoiceItems.Amount) Total
        FROM T_Invoices
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id
        WHERE Party_id= %s  and T_Invoices.InvoiceDate BETWEEN  %s AND %s  and b.GSTIN != '' and TC_InvoiceItems.GSTPercentage = 0  and a.State_id = b.State_id group by id,Description
        UNION
        SELECT 1 as id, 'Inter-State supplies to unregistered persons' Description,sum(TC_InvoiceItems.Amount) Total
        FROM T_Invoices
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id
        WHERE Party_id= %s  and T_Invoices.InvoiceDate BETWEEN %s AND %s and b.GSTIN = '' and TC_InvoiceItems.GSTPercentage = 0  and a.State_id != b.State_id group by id,Description
        UNION
        SELECT 1 as id, 'Intra-State supplies to unregistered persons' Description,sum(TC_InvoiceItems.Amount) Total
        FROM T_Invoices
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id
        WHERE Party_id=%s  and T_Invoices.InvoiceDate BETWEEN %s AND %s and b.GSTIN = '' and TC_InvoiceItems.GSTPercentage = 0  and a.State_id = b.State_id group by id,Description
        ''',([Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate]))
                # print(str(EXEMPquery.query))
                EXEMPdata = EXEMPSerializer(EXEMPquery, many=True).data
                df6=pd.DataFrame(EXEMPdata)
                ws6 = wb.create_sheet(title="EXEMP")
         
                specific_column_names = {
                'Description':'Description', 
                'Total':'NilRatedSupplies',
                'Exempted(other than nil rated/non GST supply)':'Exempted(other than nil rated/non GST supply)',
                'Non-GST Supplies':'Non-GST Supplies'
                }
                
                for col_idx, header in enumerate(specific_column_names, start=1):
                    cell = ws6.cell(row=4, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    cell.font = bold_font
                
                # Write the data on the second worksheet
                for col_idx, header in enumerate(df6.columns, start=1):
                    for row_idx, value in enumerate(df6[header], start=5):
                        ws6.cell(row=row_idx, column=col_idx, value=value)  
                
                max_cols = 4
                
                            
                ws6.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
                merged_cell = ws6.cell(row=1, column=1, value="Summary For Nil rated, exempted and non GST outward supplies (8)")
                bold_font = Font(bold=True)
                merged_cell.font = bold_font
                merged_cell.alignment = Alignment(horizontal='center')  # Align text to center
                
                
                EXEMPquery2= T_Invoices.objects.raw(''' SELECT 1 as id, '' AA,sum(A.Total) TotalNilRatedSupplies,'' TotalExemptedSupplies,'' TotalNonGSTSupplies
FROM (SELECT 1 as id , 'Inter-State supplies to registered persons' Description,sum(TC_InvoiceItems.Amount) Total,'' TotalExemptedSupplies,'' TotalNonGSTSupplies
        FROM T_Invoices
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id  
        WHERE Party_id= %s and T_Invoices.InvoiceDate BETWEEN %s AND %s and b.GSTIN != '' and TC_InvoiceItems.GSTPercentage= 0  and a.State_id != b.State_id group by id,Description
        UNION
        SELECT 1 as id, 'Intra-State supplies to registered persons' Description,sum(TC_InvoiceItems.Amount) Total,'' TotalExemptedSupplies,'' TotalNonGSTSupplies
        FROM T_Invoices
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id
        WHERE Party_id= %s  and T_Invoices.InvoiceDate BETWEEN  %s AND %s  and b.GSTIN != '' and TC_InvoiceItems.GSTPercentage = 0  and a.State_id = b.State_id group by id,Description
        UNION
        SELECT 1 as id, 'Inter-State supplies to unregistered persons' Description,sum(TC_InvoiceItems.Amount) Total,'' TotalExemptedSupplies,'' TotalNonGSTSupplies
        FROM T_Invoices
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id
        WHERE Party_id= %s  and T_Invoices.InvoiceDate BETWEEN %s AND %s and b.GSTIN = '' and TC_InvoiceItems.GSTPercentage = 0  and a.State_id != b.State_id group by id,Description
        UNION
        SELECT 1 as id, 'Intra-State supplies to unregistered persons' Description,sum(TC_InvoiceItems.Amount) Total,'' TotalExemptedSupplies,'' TotalNonGSTSupplies
        FROM T_Invoices
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_Parties a ON a.id=T_Invoices.Party_id
        JOIN M_Parties b ON b.id=T_Invoices.Customer_id
        WHERE Party_id=%s  and T_Invoices.InvoiceDate BETWEEN %s AND %s and b.GSTIN = '' and TC_InvoiceItems.GSTPercentage = 0  and a.State_id = b.State_id group by id,Description)A
        ''',([Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate]))
            
                EXEMP2data = EXEMP2Serializer2(EXEMPquery2, many=True).data
                EXEMPdf6=pd.DataFrame(EXEMP2data)
            
                specific_column_names = {
                'TotalNilRatedSupplies':'Total NilRated Supplies',
                'TotalExemptedSupplies':'Total Exempted Supplies',
                'TotalNon-GSTSupplies':'Total Non-GST Supplies',
        
                }
            
                for col_idx, header in enumerate(specific_column_names, start=2):
                    ws6.cell(row=2, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    ws6.cell(row=2, column=col_idx).font = bold_font

                for col_idx, header in enumerate(EXEMPdf6.columns, start=2):
                    for row_idx, value in enumerate(EXEMPdf6[header], start=3):
                        ws6.cell(row=row_idx, column=col_idx, value=value)    
               
                # Example data for the seven sheet HSN            
                HSNquery = T_Invoices.objects.raw('''SELECT 1 as id, M_GSTHSNCode.HSNCode,M_Items.Name Description, 'NOS-NUMBERS' AS UQC,sum(TC_InvoiceItems.QtyInNo) TotalQuantity,sum(TC_InvoiceItems.Amount)TotalValue,sum(TC_InvoiceItems.BasicAmount) TaxableValue, sum(TC_InvoiceItems.IGST)IntegratedTaxAmount,sum(TC_InvoiceItems.CGST)CentralTaxAmount,sum(TC_InvoiceItems.SGST)StateUTTaxAmount, '' CessAmount
        FROM T_Invoices 
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_GSTHSNCode ON M_GSTHSNCode.id=TC_InvoiceItems.GST_id
        JOIN M_Items ON M_Items.id=TC_InvoiceItems.Item_id
        WHERE Party_id= %s  and T_Invoices.InvoiceDate BETWEEN %s AND %s  Group by id, M_GSTHSNCode.HSNCode,M_Items.Name''',([Party],[FromDate],[ToDate]))
            
                HSNdata = HSNSerializer(HSNquery, many=True).data
                df7=pd.DataFrame(HSNdata)
                ws7 = wb.create_sheet(title="HSN")
               
                specific_column_names = {
                'HSNCode':'HSN', 
                'Description':'Description',
                'UQC':'UQC',
                'TotalQuantity':'TotalQuantity',
                'TotalValue':'Total Value',
                'TaxableValue' :'Taxable Value',
                'IntegratedTaxAmount' :'Integrated Tax Amount',
                'CentralTaxAmount' :'Central Tax Amount',
                'StateUTTaxAmount' :'State/UT Tax Amount',
                'CessAmount':'Cess Amount',
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    cell = ws7.cell(row=4, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    cell.font = bold_font
                
                # Write the data on the second worksheet
                for col_idx, header in enumerate(df7.columns, start=1):
                    for row_idx, value in enumerate(df7[header], start=5):
                        ws7.cell(row=row_idx, column=col_idx, value=value)                        
            
                max_cols = 10
                
                ws7.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
                merged_cell = ws7.cell(row=1, column=1, value="Summary For HSN(12)")
                bold_font = Font(bold=True)
                merged_cell.font = bold_font
                merged_cell.alignment = Alignment(horizontal='center')  # Align text to center
            
                HSNquery2= T_Invoices.objects.raw('''SELECT 1 as id, COUNT(DISTINCT(A.HSNCode))NoofHSN,''a,''b,''c,sum(A.TotalValue) TotalValue,sum(A.TaxableValue) TaxableValue,sum(A.IntegratedTaxAmount) IntegratedTaxAmount,sum(A.CentralTaxAmount) CentralTaxAmount,sum(A.StateUTTaxAmount) StateUTTaxAmount, '' CessAmount
FROM (SELECT 1 as id, M_GSTHSNCode.HSNCode,M_Items.Name Description, 'NOS-NUMBERS' AS UQC,sum(TC_InvoiceItems.QtyInNo) TotalQuantity,sum(TC_InvoiceItems.Amount)TotalValue,sum(TC_InvoiceItems.BasicAmount) TaxableValue, sum(TC_InvoiceItems.IGST)IntegratedTaxAmount,sum(TC_InvoiceItems.CGST)CentralTaxAmount,sum(TC_InvoiceItems.SGST)StateUTTaxAmount, '' CessAmount
        FROM T_Invoices 
        JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
        JOIN M_GSTHSNCode ON M_GSTHSNCode.id=TC_InvoiceItems.GST_id
        JOIN M_Items ON M_Items.id=TC_InvoiceItems.Item_id
        WHERE Party_id=%s  and T_Invoices.InvoiceDate BETWEEN  %s AND  %s  Group by id, M_GSTHSNCode.HSNCode,M_Items.Name)A''',([Party],[FromDate],[ToDate]))
            
                HSN2data = HSN2Serializer2(HSNquery2, many=True).data
                HSNdf7=pd.DataFrame(HSN2data)
            
                specific_column_names = {
                'NoofHSN':'No. of HSN', 
                'a':'', 
                'b':'', 
                'c':'', 
                'TotalValue':'Total Value',
                'TaxableValue' :'Total Taxable Value',
                'IntegratedTaxAmount' :'Total Integrated Tax Amount',
                'CentralTaxAmount' :'Total Central Tax Amount',
                'StateUTTaxAmount' :'Total State/UT Tax Amount',
                'CessAmount':'Total Cess Amount',
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    ws7.cell(row=2, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    ws7.cell(row=2, column=col_idx).font = bold_font

                for col_idx, header in enumerate(HSNdf7.columns, start=1):
                    for row_idx, value in enumerate(HSNdf7[header], start=3):
                        ws7.cell(row=row_idx, column=col_idx, value=value)
                 
            
                
                # Example data for the eight sheet Docs         
                Docsquery = T_Invoices.objects.raw('''SELECT 1 as id, 'Invoices for outward supply' a,MIN(T_Invoices.InvoiceNumber)MINID,max(T_Invoices.InvoiceNumber)MAXID ,count(*)cnt,(SELECT count(*)cnt from T_DeletedInvoices  where Party =%s and T_DeletedInvoices.InvoiceDate BETWEEN %s AND %s ) Cancelledcnt ,'1' b
                FROM T_Invoices  where Party_id =%s and T_Invoices.InvoiceDate BETWEEN %s AND %s
                UNION 
                SELECT 1 as id, 'Credit Note' a,MIN(T_CreditDebitNotes.FullNoteNumber)MINID,MAX(T_CreditDebitNotes.FullNoteNumber)MAXID ,count(*)cnt,'0' Cancelledcnt ,'2' b
                FROM T_CreditDebitNotes 
                WHERE  T_CreditDebitNotes.NoteType_id=37 and Party_id =%s and T_CreditDebitNotes.CRDRNoteDate BETWEEN %s AND %s
                UNION 
                SELECT 1 as id, 'Debit Note' a, MIN(T_CreditDebitNotes.FullNoteNumber)MINID,MAX(T_CreditDebitNotes.FullNoteNumber)MAXID ,count(*)cnt,'0' Cancelledcnt,'3' b 
                FROM T_CreditDebitNotes  
                WHERE  T_CreditDebitNotes.NoteType_id=38 AND Party_id =%s and T_CreditDebitNotes.CRDRNoteDate BETWEEN %s AND %s ''',([Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate]))
                   
                Docsdata = DocsSerializer(Docsquery, many=True).data
                df8=pd.DataFrame(Docsdata)
                ws8 = wb.create_sheet(title="Docs")
                
                    
                specific_column_names = {
                'a':'Nature of Document', 
                'MINID':'Sr. No. From',
                'MAXID':'Sr. No. To',
                'cnt':'Total Number',
                'Cancelledcnt':'Cancelled'
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    cell = ws8.cell(row=4, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    cell.font = bold_font
                
                # Write the data on the second worksheet
                for col_idx, header in enumerate(df8.columns, start=1):
                    for row_idx, value in enumerate(df8[header], start=5):
                        ws8.cell(row=row_idx, column=col_idx, value=value)                
                    
                max_cols = 5
                
                ws8.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
                merged_cell = ws8.cell(row=1, column=1, value="Summary of documents issued during the tax period (13)")
                bold_font = Font(bold=True)
                merged_cell.font = bold_font
                merged_cell.alignment = Alignment(horizontal='center')  # Align text to center
                
                
                Docsquery2 = T_Invoices.objects.raw(''' SELECT 1 as id, '' AA,'' bb, '' cc,sum(A.cnt)cnt ,sum(A.Cancelledcnt)Cancelledcnt
FROM (SELECT 1 as id, 'Invoices for outward supply' a,MIN(T_Invoices.InvoiceNumber)MINID,max(T_Invoices.InvoiceNumber)MAXID ,count(*)cnt,(SELECT count(*)cnt from T_DeletedInvoices  where Party =%s and T_DeletedInvoices.InvoiceDate BETWEEN %s AND %s ) Cancelledcnt ,'1' b
                FROM T_Invoices  where Party_id =%s and T_Invoices.InvoiceDate BETWEEN %s AND %s
                UNION 
                SELECT 1 as id, 'Credit Note' a,MIN(T_CreditDebitNotes.FullNoteNumber)MINID,MAX(T_CreditDebitNotes.FullNoteNumber)MAXID ,count(*)cnt,'0' Cancelledcnt ,'2' b
                FROM T_CreditDebitNotes 
                WHERE  T_CreditDebitNotes.NoteType_id=37 and Party_id =%s and T_CreditDebitNotes.CRDRNoteDate BETWEEN %s AND %s
                UNION 
                SELECT 1 as id, 'Debit Note' a, MIN(T_CreditDebitNotes.FullNoteNumber)MINID,MAX(T_CreditDebitNotes.FullNoteNumber)MAXID ,count(*)cnt,'0' Cancelledcnt,'3' b 
                FROM T_CreditDebitNotes  
                WHERE  T_CreditDebitNotes.NoteType_id=38 AND Party_id =%s and T_CreditDebitNotes.CRDRNoteDate BETWEEN %s AND %s )A ''',([Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate])) 
            
                Docsdata2 = Docs2Serializer2(Docsquery2, many=True).data
                Docsdf8=pd.DataFrame(Docsdata2)
            
                specific_column_names = {
                    'AA':'', 
                    'bb':'', 
                    'cc':'', 
                    'cnt':'Total Numbers',
                    'Cancelledcnt':'Total Cancelled'
                }
            
                for col_idx, header in enumerate(specific_column_names, start=1):
                    ws8.cell(row=2, column=col_idx, value=specific_column_names.get(header, header))
                    bold_font = Font(bold=True)
                    ws8.cell(row=2, column=col_idx).font = bold_font

                for col_idx, header in enumerate(Docsdf8.columns, start=1):
                    for row_idx, value in enumerate(Docsdf8[header], start=3):
                        ws8.cell(row=row_idx, column=col_idx, value=value)
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=example.xlsx'

                return response
        except Exception as e:
            return JsonResponse({'StatusCode': 400, 'Status': True, 'Message':  Exception(e), 'Data': []})

