from decimal import Decimal
from django.http import JsonResponse
from rest_framework.generics import CreateAPIView
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from rest_framework.parsers import JSONParser
from ..Serializer.S_R3B_Reports import *

from ..models import  *
import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
import pandas as pd



class GSTR3BDownloadView(CreateAPIView):
    permission_classes = (IsAuthenticated,)
   
    def post(self, request):
        Orderdata = JSONParser().parse(request)
        FromDate = Orderdata['FromDate']
        ToDate = Orderdata['ToDate']
        Party = Orderdata['Party']
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '3.1'
        query = T_Invoices.objects.raw('''SELECT 1 as id, '(a) Outward Taxable  supplies  (other than zero rated, nil rated and exempted)' A,sum(TC_InvoiceItems.BasicAmount) Taxablevalue,SUM(TC_InvoiceItems.IGST) IGST,sum(TC_InvoiceItems.CGST) CGST, SUM(TC_InvoiceItems.SGST)SGST,'0'Cess
FROM T_Invoices
JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
WHERE Party_id=%s AND  T_Invoices.InvoiceDate BETWEEN %s AND %s and  TC_InvoiceItems.GSTPercentage != 0
UNION 
SELECT 1 as id, '(b) Outward Taxable  supplies  (zero rated )' A, 0 Taxablevalue,0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id,'(c) Other Outward Taxable  supplies (Nil rated, exempted)' A,sum(TC_InvoiceItems.BasicAmount) Taxablevalue,SUM(TC_InvoiceItems.IGST) IGST,sum(TC_InvoiceItems.CGST) CGST, SUM(TC_InvoiceItems.SGST)SGST,'0'Cess
FROM T_Invoices
JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
WHERE  Party_id=%s AND T_Invoices.InvoiceDate BETWEEN  %s AND %s AND TC_InvoiceItems.GSTPercentage = 0
UNION 
SELECT 1 as id, '(d) Inward supplies (liable to reverse charge) ' A, 0 Taxablevalue,0 IGST,0 CGST, 0 SGST,0 Cess
UNION 
SELECT 1 as id, '(e) Non-GST Outward supplies' A, 0 Taxablevalue,0 IGST,0 CGST, 0 SGST,0 Cess''',([Party],[FromDate],[ToDate],[Party],[FromDate],[ToDate]))
        
        DOSAISLTRCdata = DOSAISLTRCSerializer(query, many=True).data
        df=pd.DataFrame(DOSAISLTRCdata)
        if not df.empty:
        # print (df)
            specific_column_names = {
            'A':'Nature of Supplies', 
            'Taxablevalue':'Total Taxable value',
            'IGST':'Integrated Tax',
            'CGST':'Central Tax',
            'SGST':'State/UT Tax',
            'Cess':'Cess',
            }
        
            # Define which columns header Font bold
            for col_idx, header in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=specific_column_names.get(header, header))
                bold_font = Font(bold=True)
                cell.font = bold_font
                
            # Write the data
            for col_idx, header in enumerate(df.columns, start=1):
                for row_idx, value in enumerate(df[header], start=2):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                
################################## 4.EligibleITC #############################################################################

        EligibleITCquery = T_Invoices.objects.raw('''SELECT 1 as id, '(A) ITC Available (Whether in full or part)' A,0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id,  '(1)   Import of goods ' A, 0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id, '(2)   Import of services' A,0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id, '(3)   Inward supplies liable to reverse charge(other than 1 &2 above)' A, 0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id, '(4)   Inward supplies from ISD' A, 0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id, '(5)   All other ITC' A,IFNULL(SUM(TC_InvoiceItems.IGST),0) IGST,IFNULL(sum(TC_InvoiceItems.CGST),0) CGST, IFNULL(SUM(TC_InvoiceItems.SGST),0) SGST,'0'Cess
FROM T_Invoices
JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
WHERE  T_Invoices.Customer_id=%s AND T_Invoices.InvoiceDate BETWEEN %s AND %s
UNION
SELECT 1 as id, ' (B) ITC Reversed' A, 0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id, '(1)   As per Rule 42 & 43 of SGST/CGST rules ' A, 0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id, '(2)   Others' A, 0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id, '(C) Net ITC Available (A)-(B)' A, 0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id, ' (D) Ineligible ITC' A, 0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id, '(1)   As per section 17(5) of CGST//SGST Act' A, 0 IGST,0 CGST, 0 SGST,0 Cess
UNION
SELECT 1 as id, '(2)   Others' A, 0 IGST,0 CGST, 0 SGST,0 Cess''',([Party],[FromDate],[ToDate]))
        EgibleITCdata = EligibleITCSerializer(EligibleITCquery, many=True).data
        df2=pd.DataFrame(EgibleITCdata)
        ws2 = wb.create_sheet(title="Eligible ITC")
        if not df2.empty:
            
            specific_column_names = {
            'A':'Details', 
            'IGST':'Integrated Tax',
            'CGST':'Central Tax',
            'SGST':'State/UT Tax',
            'Cess':'Cess',
            }
        
            for col_idx, header in enumerate(df2.columns, start=1):
                cell = ws2.cell(row=1, column=col_idx, value=specific_column_names.get(header, header))
                bold_font = Font(bold=True)
                cell.font = bold_font
            
            # Write the data on the second worksheet
            for col_idx, header in enumerate(df2.columns, start=1):
                for row_idx, value in enumerate(df2[header], start=2):
                    ws2.cell(row=row_idx, column=col_idx, value=value) 
        
        
#####################   3.2 Of the supplies shown in 3.1 (a), details of inter-state supplies made to unregistered persons, composition taxable person and UIN holders#################################################
        
        query3 = T_Invoices.objects.raw('''SELECT 1 as id, concat(M_States.StateCode,'-',M_States.Name)states ,sum(TC_InvoiceItems.BasicAmount) Taxablevalue,SUM(TC_InvoiceItems.IGST) IGST
FROM T_Invoices
JOIN TC_InvoiceItems ON TC_InvoiceItems.Invoice_id=T_Invoices.id
JOIN M_Parties ON M_Parties.id=T_Invoices.Customer_id
JOIN M_States ON M_Parties.State_id=M_States.id
WHERE  Party_id=%s AND T_Invoices.InvoiceDate BETWEEN %s AND %s   group by M_States.id''',([Party],[FromDate],[ToDate]))
        Query3data = Query3Serializer(query3, many=True).data
        df3=pd.DataFrame(Query3data)
        ws3 = wb.create_sheet(title="3.2")
        if not df3.empty:

            specific_column_names = {
            'states':'Place of Supply(State/UT)', 
            'Taxablevalue' :'Taxable Value',
            'IGST':'Amount of Integrated Tax',
            }
        
            for col_idx, header in enumerate(df3.columns, start=1):
                cell = ws3.cell(row=1, column=col_idx, value=specific_column_names.get(header, header))
                bold_font = Font(bold=True)
                cell.font = bold_font
            
            # Write the data on the second worksheet
            for col_idx, header in enumerate(df3.columns, start=1):
                for row_idx, value in enumerate(df3[header], start=2):
                    ws3.cell(row=row_idx, column=col_idx, value=value) 


        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=example.xlsx'

        return response